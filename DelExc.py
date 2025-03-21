import openpyxl
import math
from tqdm import tqdm
print("Скрипт пока еще только учится, и делить на равные части будет содержимое лишь первого листа.")

def divide_excel_file(input_file, num_files):
    # Загрузка входного файла Excel
    workbook = openpyxl.load_workbook(input_file)
    sheet = workbook.active

    # Подсчет количества строк в файле
    num_lines = sheet.max_row

    # Вычисление количества строк на файл
    lines_per_file = math.ceil(num_lines / num_files)

    # Создание выходных файлов и деление строк
    current_file = 1
    current_line = 1

    pbar = tqdm(total=num_lines, desc="Обработка строк")

    while current_line <= num_lines:
        # Создание новой рабочей книги для каждого выходного файла
        output_workbook = openpyxl.Workbook()
        output_sheet = output_workbook.active

        # Копирование строк на выходной лист
        for _ in range(lines_per_file):
            if current_line > num_lines:
                break

            for col in range(1, sheet.max_column + 1):
                output_sheet.cell(row=(current_line - 1) % lines_per_file + 1,
                                  column=col).value = sheet.cell(
                    row=current_line, column=col).value

            current_line += 1
            pbar.update(1)

        # Сохранение выходного файла
        output_filename = f"output_{current_file}.xlsx"
        output_workbook.save(output_filename)
        current_file += 1

    pbar.close()

    print(f"{num_lines} строк успешно разделены на {num_files} файлов!")


# Пример использования
input_file = str(input(r"Введите путь до файла с его разрешением .xlsx: ")) #r"C:\Users\1\Desktop\excel\28_ 12-59-56.xlsx"
num_files = int(input("Введите количество выходных файлов: "))
divide_excel_file(input_file, num_files)